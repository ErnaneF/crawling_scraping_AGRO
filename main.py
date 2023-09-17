from bs4 import BeautifulSoup
import requests
import pandas as pd  # Alterei para 'pd' para facilitar a referência
from openpyxl import Workbook

headers = {
    'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36"
}
texto_html = 'https://www.soybeansandcorn.com/articles/'

def conteudo(pag_num):
    url = f'{texto_html}?page={pag_num}&'
    resposta = requests.get(url, headers=headers)  # Adicionei headers aqui
    html = resposta.content
    soup = BeautifulSoup(html, 'html.parser')
    noticias = soup.find_all('li', class_=None)
    conjunto = []
    for noticia in noticias:
        texto_noticia = noticia.find('a').text
        antes, depois = texto_noticia.split(":", 1)
        noticia_inteira = noticia.find('a')['href']

        conjunto.append({'Data': antes, 'Titulo': depois, 'Noticia Inteira': noticia_inteira})  # Corrigi os nomes das colunas
    return conjunto

pag_num = 1
tdsD = []

while True:
    pag = conteudo(pag_num)
    if not pag:
        break
    tdsD.extend(pag)
    pag_num += 1

data = pd.DataFrame(tdsD)

excel = "Noticias.xlsx"
data.to_excel(excel, index=False, engine='openpyxl')

wb = Workbook()
ws = wb.active
ws.title = "Not"

ws['A1'] = 'Data'
ws['B1'] = 'Titulo'  # Corrigi o nome da coluna
ws['C1'] = 'Noticia Inteira'

for row_index, row_data in data.iterrows():
    ws.cell(row=row_index+2, column=1, value=row_data['Data'])
    ws.cell(row=row_index+2, column=2, value=row_data['Titulo'])  # Corrigi o nome da coluna
    ws.cell(row=row_index+2, column=3, value=row_data['Noticia Inteira'])  # Corrigi o nome da coluna
wb.save(excel)
print(f"Dados salvos em '{excel}'.")





















''''
def prox_pag(soup):
    paginas = soup.find('a', {'class': 'page link'}) #prox pág
    if not paginas.find('li', {'class': 'page-item next disabled'}):
        url = 'https://www.soybeansandcorn.com/'
        prox = soup.find('a','page link', href = True)
        ult_url = (url + str(prox['href']))
        return ult_url
    else:
        return

conteudo(soup)'''